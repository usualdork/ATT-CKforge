import requests
import json
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import os
import sys
from datetime import datetime

class MitreAttackMatrixFetcher:
    """
    A class to fetch and process MITRE ATT&CK matrices interactively for any platform.
    """
    
    def __init__(self):
        self.urls = {
            "enterprise": "https://raw.githubusercontent.com/mitre/cti/master/enterprise-attack/enterprise-attack.json",
            "mobile": "https://raw.githubusercontent.com/mitre/cti/master/mobile-attack/mobile-attack.json",
            "ics": "https://raw.githubusercontent.com/mitre/cti/master/ics-attack/ics-attack.json"
        }
        self.output_dir = "mitre_matrices"
        
        # Create output directory if it doesn't exist
        if not os.path.exists(self.output_dir):
            print(f"Creating output directory: {self.output_dir}")
            os.makedirs(self.output_dir)
        
        # Define styles for Excel formatting
        self.header_font = Font(bold=True)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        self.tactic_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
        
        # Cache for attack data
        self.attack_data = {}
    
    def fetch_mitre_data(self, matrix_type):
        """Fetch MITRE ATT&CK data for a specific matrix type"""
        if matrix_type in self.attack_data:
            return self.attack_data[matrix_type]
            
        url = self.urls.get(matrix_type)
        if not url:
            print(f"Error: Unknown matrix type '{matrix_type}'")
            return None
            
        print(f"Fetching MITRE ATT&CK data for {matrix_type} from {url}...")
        try:
            response = requests.get(url)
            response.raise_for_status()
            data = response.json()
            self.attack_data[matrix_type] = data
            return data
        except Exception as e:
            print(f"Error fetching MITRE data: {e}")
            return None
    
    def extract_platforms(self, data):
        """Extract available platforms from the MITRE data"""
        platforms = set()
        for obj in data['objects']:
            if obj.get('type') == 'attack-pattern' and 'x_mitre_platforms' in obj:
                for platform in obj['x_mitre_platforms']:
                    platforms.add(platform)
        return sorted(list(platforms))
    
    def extract_tactics(self, data):
        """Extract all tactics (kill chain phases) from the MITRE data"""
        tactics = {}
        for obj in data['objects']:
            if obj.get('type') == 'x-mitre-tactic':
                shortname = obj.get('x_mitre_shortname', '')
                if shortname:
                    tactics[shortname] = {
                        'name': obj.get('name', ''),
                        'id': obj.get('external_references', [{}])[0].get('external_id', '') if obj.get('external_references') else '',
                        'url': obj.get('external_references', [{}])[0].get('url', '') if obj.get('external_references') else '',
                    }
        return tactics
    
    def build_technique_dict(self, data):
        """Build a dictionary of techniques with their subtechniques"""
        techniques = {}
        for obj in data['objects']:
            if obj.get('type') == 'attack-pattern':
                if not obj.get('external_references'):
                    continue
                    
                technique_id = obj['external_references'][0].get('external_id', '')
                if not technique_id:
                    continue
                    
                techniques[technique_id] = {
                    'id': technique_id,
                    'name': obj.get('name', ''),
                    'platforms': obj.get('x_mitre_platforms', []),
                    'tactics': obj.get('kill_chain_phases', []),
                    'url': obj['external_references'][0].get('url', ''),
                    'is_subtechnique': '.' in technique_id,
                    'parent_technique': technique_id.split('.')[0] if '.' in technique_id else None,
                    'subtechniques': []
                }
        
        # Link subtechniques to their parent techniques
        for tech_id, tech in list(techniques.items()):
            if tech['is_subtechnique'] and tech['parent_technique'] in techniques:
                parent = techniques[tech['parent_technique']]
                parent['subtechniques'].append(tech_id)
                
        return techniques
    
    def build_matrix_for_platform(self, data, platform):
        """Build a matrix for a specific platform"""
        print(f"Building matrix for {platform}...")
        tactics = self.extract_tactics(data)
        techniques = self.build_technique_dict(data)
        
        # Create a matrix organized by tactics
        matrix = {}
        for tech_id, tech in techniques.items():
            # Skip subtechniques as they will be processed with parent techniques
            if tech['is_subtechnique']:
                continue
                
            # Only include techniques available for this platform
            if platform not in tech['platforms']:
                continue
                
            # Add the technique to each tactic it belongs to
            for phase in tech.get('tactics', []):
                tactic_shortname = phase.get('phase_name', '')
                if not tactic_shortname or tactic_shortname not in tactics:
                    continue
                    
                if tactic_shortname not in matrix:
                    matrix[tactic_shortname] = []
                    
                # Only add subtechniques for this platform
                valid_subtechniques = []
                for sub_id in tech['subtechniques']:
                    if platform in techniques[sub_id]['platforms']:
                        valid_subtechniques.append({
                            'name': techniques[sub_id]['name'],
                            'reference': techniques[sub_id]['url']
                        })
                        
                matrix[tactic_shortname].append({
                    'technique': tech['name'],
                    'reference': tech['url'],
                    'subtechniques': valid_subtechniques
                })
        
        return matrix, tactics
    
    def create_excel_from_matrix(self, matrix, tactics, platform, matrix_type):
        """Create a formatted Excel file from the matrix"""
        file_name = os.path.join(self.output_dir, f"MITRE_ATT&CK_{matrix_type}_{platform}_{datetime.now().strftime('%Y%m%d')}.xlsx")
        print(f"Creating Excel file: {file_name}")
        
        # Create new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"ATT&CK Matrix - {platform}"
        
        # Apply column widths
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 35
        
        row = 1
        # Add a title for the workbook
        ws.cell(row=row, column=1).value = f"MITRE ATT&CK {matrix_type.title()} Matrix for {platform}"
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        row += 2  # Add space after title
        
        # Track if any tactics were processed
        tactics_processed = False
        
        # Process each tactic in the matrix
        for tactic_shortname, techniques in matrix.items():
            if not techniques:  # Skip empty tactics
                continue
                
            tactics_processed = True
                
            # Add tactic header
            tactic_name = tactics[tactic_shortname]['name']
            tactic_cell = ws.cell(row=row, column=1)
            tactic_cell.value = tactic_name
            tactic_cell.font = self.header_font
            tactic_cell.fill = self.tactic_fill
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            tactic_cell.alignment = Alignment(horizontal='left')
            row += 1
            
            # Add column headers
            headers = ["Technique", "Reference", "Subtechnique", "Subtechnique Reference"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = header
                cell.font = self.header_font
                cell.border = self.border
                cell.fill = self.header_fill
            row += 1
            
            # Track start row for this tactic's techniques
            tactic_start_row = row
            
            # Add techniques and subtechniques
            for technique in techniques:
                if not technique['subtechniques']:
                    # Write just the technique with no subtechniques
                    ws.cell(row=row, column=1).value = technique['technique']
                    ws.cell(row=row, column=2).value = technique['reference']
                    row += 1
                else:
                    # Write technique with first subtechnique
                    ws.cell(row=row, column=1).value = technique['technique']
                    ws.cell(row=row, column=2).value = technique['reference']
                    ws.cell(row=row, column=3).value = technique['subtechniques'][0]['name']
                    ws.cell(row=row, column=4).value = technique['subtechniques'][0]['reference']
                    row += 1
                    
                    # Write remaining subtechniques with empty technique cells
                    for sub in technique['subtechniques'][1:]:
                        ws.cell(row=row, column=3).value = sub['name']
                        ws.cell(row=row, column=4).value = sub['reference']
                        row += 1
            
            # Apply borders to all cells in the tactic section
            for r in range(tactic_start_row, row):
                for c in range(1, 5):
                    ws.cell(row=r, column=c).border = self.border
            
            # Add empty rows between tactics
            row += 2
        
        if not tactics_processed:
            ws.cell(row=row, column=1).value = "No techniques found for this platform in the framework."
            return None
            
        # Save the workbook
        try:
            wb.save(file_name)
            print(f"Successfully created {file_name}")
            return file_name
        except Exception as e:
            print(f"Error saving Excel file: {e}")
            return None
    
    def process_selection(self, matrix_type, platform):
        """Process a user-selected matrix and platform"""
        data = self.fetch_mitre_data(matrix_type)
        if not data:
            return None
            
        matrix, tactics = self.build_matrix_for_platform(data, platform)
        return self.create_excel_from_matrix(matrix, tactics, platform, matrix_type)


def clear_screen():
    """Clear the console screen"""
    os.system('cls' if os.name == 'nt' else 'clear')


def display_menu(options, title, multi_select=False):
    """Display a menu and get user selection(s)"""
    clear_screen()
    print(f"\n=== {title} ===\n")
    
    for i, option in enumerate(options, 1):
        print(f"{i}. {option}")
    
    if multi_select:
        print("\nEnter numbers separated by commas (e.g., 1,3,5)")
        print("Or enter 'all' to select all options")
        print("Or enter 'q' to quit")
        
        while True:
            selection = input("\nYour selection: ").strip().lower()
            
            if selection == 'q':
                return []
                
            if selection == 'all':
                return list(range(len(options)))
                
            try:
                # Parse comma-separated values
                selections = [int(x.strip()) - 1 for x in selection.split(',')]
                # Validate selections
                if all(0 <= s < len(options) for s in selections):
                    return selections
                else:
                    print("Invalid selection. Please try again.")
            except ValueError:
                print("Invalid input. Please enter numbers separated by commas.")
    else:
        print("\nEnter a number to select an option")
        print("Or enter 'q' to quit")
        
        while True:
            selection = input("\nYour selection: ").strip().lower()
            
            if selection == 'q':
                return -1
                
            try:
                selected = int(selection) - 1
                if 0 <= selected < len(options):
                    return selected
                else:
                    print("Invalid selection. Please try again.")
            except ValueError:
                print("Invalid input. Please enter a number.")


def interactive_menu():
    """Run the interactive menu system"""
    fetcher = MitreAttackMatrixFetcher()
    
    # Main matrix type selection
    matrix_types = ["Enterprise", "Mobile", "ICS"]
    selected = display_menu(matrix_types, "Select MITRE ATT&CK Matrix Type")
    
    if selected == -1:
        return
    
    matrix_type = matrix_types[selected].lower()
    
    # Fetch data for the selected matrix
    data = fetcher.fetch_mitre_data(matrix_type)
    if not data:
        print("Failed to fetch data. Exiting.")
        return
        
    # Get available platforms
    platforms = fetcher.extract_platforms(data)
    
    if not platforms:
        print("No platforms found in the selected matrix.")
        return
        
    # Platform selection (multi-select)
    platform_indices = display_menu(platforms, f"Select Platforms ({matrix_type.title()})", multi_select=True)
    
    if not platform_indices:
        print("No platforms selected. Exiting.")
        return
        
    # Process each selected platform
    created_files = []
    for idx in platform_indices:
        platform = platforms[idx]
        file_path = fetcher.process_selection(matrix_type, platform)
        if file_path:
            created_files.append(file_path)
            
    # Display results
    print("\nProcessing complete!")
    if created_files:
        print("\nCreated files:")
        for file in created_files:
            print(f" - {file}")
    else:
        print("No files were created.")
        
    input("\nPress Enter to continue...")


def main():
    """Main function"""
    try:
        while True:
            clear_screen()
            print("\n=== MITRE ATT&CK Matrix Generator ===\n")
            print("1. Start Interactive Selection")
            print("2. Exit")
            
            choice = input("\nEnter your choice (1-2): ")
            
            if choice == '1':
                interactive_menu()
            elif choice == '2':
                print("Exiting program. Goodbye!")
                break
            else:
                print("Invalid choice. Press Enter to try again.")
                input()
                
    except KeyboardInterrupt:
        print("\nProgram terminated by user.")
    except Exception as e:
        print(f"\nAn error occurred: {e}")
        input("Press Enter to exit...")


if __name__ == "__main__":
    main()
